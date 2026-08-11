import unittest

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.formula import ArrayFormula

from generic_grader.excel.formulas_match_reference import build
from generic_grader.utils.options import Options


def write_workbook(path, sheet_name="Sheet1", cells=None):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    for cell, value in (cells or {}).items():
        worksheet[cell] = value
    workbook.save(path)


@pytest.fixture()
def built_class():
    return build(
        Options(
            required_files=("submission.xlsx",),
            entries=("A1", "B1"),
            kwargs={"reference_file": "reference.xlsx"},
        )
    )


@pytest.fixture()
def built_instance(built_class):
    return built_class()


def test_build_class(built_class):
    assert issubclass(built_class, unittest.TestCase)


def test_build_class_name(built_class):
    assert built_class.__name__ == "TestFormulasMatchReference"


def test_instance_has_test_method(built_instance):
    assert hasattr(built_instance, "test_formulas_match_reference_0")


def test_doc_func_same_range_default(built_instance):
    assert (
        built_instance.test_formulas_match_reference_0.__doc__
        == "Check that formulas in range A1:B1 on sheet `<first worksheet>` exactly match the reference formulas."
    )


def test_doc_func_search_mode():
    built_class = build(
        Options(
            required_files=("submission.xlsx",),
            entries=("A1", "B1"),
            range_matches_reference=False,
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class()

    assert (
        built_instance.test_formulas_match_reference_0.__doc__
        == "Check that formulas in range A1:B1 on sheet `<first worksheet>` match the reference formulas in a same-sized region somewhere in the submission workbook."
    )


def test_passing_same_range_case(fix_syspath):
    write_workbook(
        fix_syspath / "reference.xlsx",
        cells={"A1": "=A3+B3", "B1": "=A4+B4"},
    )
    write_workbook(
        fix_syspath / "submission.xlsx",
        cells={"A1": "=A3+B3", "B1": "=A4+B4"},
    )

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A1", "B1"),
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_formulas_match_reference_0")
    test_method = built_instance.test_formulas_match_reference_0
    test_method()

    assert test_method.__score__ == test_method.__weight__


def test_failing_same_range_case(fix_syspath):
    write_workbook(
        fix_syspath / "reference.xlsx",
        cells={"A1": "=A3+B3", "B1": "=A4+B4"},
    )
    write_workbook(
        fix_syspath / "submission.xlsx",
        cells={"A1": "=A3+B3", "B1": "=A4-B4"},
    )

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A1", "B1"),
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_formulas_match_reference_0")
    test_method = built_instance.test_formulas_match_reference_0

    with pytest.raises(AssertionError) as exc_info:
        test_method()

    assert "cell B1" in str(exc_info.value)
    assert test_method.__score__ == 0


def test_passing_search_whole_sheet_case(fix_syspath):
    write_workbook(
        fix_syspath / "reference.xlsx",
        cells={"A1": "=A3+B3", "B1": "=A4+B4"},
    )
    write_workbook(
        fix_syspath / "submission.xlsx",
        cells={"D6": "=A3+B3", "E6": "=A4+B4"},
    )

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A1", "B1"),
            range_matches_reference=False,
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_formulas_match_reference_0")
    test_method = built_instance.test_formulas_match_reference_0
    test_method()

    assert test_method.__score__ == test_method.__weight__


def test_failing_search_whole_sheet_case(fix_syspath):
    write_workbook(
        fix_syspath / "reference.xlsx",
        cells={"A1": "=A3+B3", "B1": "=A4+B4"},
    )
    write_workbook(
        fix_syspath / "submission.xlsx",
        cells={"D6": "=A3+B3", "E6": "=A4-B4"},
    )

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A1", "B1"),
            range_matches_reference=False,
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_formulas_match_reference_0")
    test_method = built_instance.test_formulas_match_reference_0

    with pytest.raises(AssertionError) as exc_info:
        test_method()

    assert "Could not find a" in str(exc_info.value)
    assert test_method.__score__ == 0


def test_first_worksheet_default_sheet_fallback(fix_syspath):
    write_workbook(
        fix_syspath / "reference.xlsx",
        sheet_name="Grades",
        cells={"A1": "=A3+B3", "B1": "=A4+B4"},
    )
    write_workbook(
        fix_syspath / "submission.xlsx",
        sheet_name="Grades",
        cells={"A1": "=A3+B3", "B1": "=A4+B4"},
    )

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A1", "B1"),
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_formulas_match_reference_0")
    test_method = built_instance.test_formulas_match_reference_0
    test_method()

    assert test_method.__score__ == test_method.__weight__


# ---------------------------------------------------------------------------
# Formula-detection/comparison gate: ordinary formula, array formula, plain
# value, and empty cell. Array formulas are openpyxl `ArrayFormula` objects
# with no `__eq__`, so two identical array formulas previously never
# compared equal -- this is the bug these tests pin down.
# ---------------------------------------------------------------------------


def test_ordinary_string_formula_matches(fix_syspath):
    write_workbook(fix_syspath / "reference.xlsx", cells={"A1": "=1+1"})
    write_workbook(fix_syspath / "submission.xlsx", cells={"A1": "=1+1"})

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A1", "A1"),
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_formulas_match_reference_0")
    test_method = built_instance.test_formulas_match_reference_0
    test_method()

    assert test_method.__score__ == test_method.__weight__


def test_identical_array_formulas_match(fix_syspath):
    for name in ("reference.xlsx", "submission.xlsx"):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"
        worksheet["A1"] = ArrayFormula("A1", "=MODE.MULT(B1:B10)")
        workbook.save(fix_syspath / name)

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A1", "A1"),
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_formulas_match_reference_0")
    test_method = built_instance.test_formulas_match_reference_0
    test_method()

    assert test_method.__score__ == test_method.__weight__


def test_different_array_formulas_do_not_match(fix_syspath):
    ref_workbook = Workbook()
    ref_worksheet = ref_workbook.active
    ref_worksheet.title = "Sheet1"
    ref_worksheet["A1"] = ArrayFormula("A1", "=MODE.MULT(B1:B10)")
    ref_workbook.save(fix_syspath / "reference.xlsx")

    sub_workbook = Workbook()
    sub_worksheet = sub_workbook.active
    sub_worksheet.title = "Sheet1"
    sub_worksheet["A1"] = ArrayFormula("A1", "=MODE.MULT(B1:B20)")
    sub_workbook.save(fix_syspath / "submission.xlsx")

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A1", "A1"),
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_formulas_match_reference_0")
    test_method = built_instance.test_formulas_match_reference_0

    with pytest.raises(AssertionError) as exc_info:
        test_method()

    assert "cell A1" in str(exc_info.value)
    assert test_method.__score__ == 0


def test_matching_plain_values_match(fix_syspath):
    write_workbook(fix_syspath / "reference.xlsx", cells={"A1": 42})
    write_workbook(fix_syspath / "submission.xlsx", cells={"A1": 42})

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A1", "A1"),
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_formulas_match_reference_0")
    test_method = built_instance.test_formulas_match_reference_0
    test_method()

    assert test_method.__score__ == test_method.__weight__


def test_matching_empty_cells_match(fix_syspath):
    write_workbook(fix_syspath / "reference.xlsx", cells={})
    write_workbook(fix_syspath / "submission.xlsx", cells={})

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A1", "A1"),
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_formulas_match_reference_0")
    test_method = built_instance.test_formulas_match_reference_0
    test_method()

    assert test_method.__score__ == test_method.__weight__
