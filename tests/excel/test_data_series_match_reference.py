import unittest

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.formula import ArrayFormula

from generic_grader.excel.data_series_match_reference import build
from generic_grader.utils.options import Options


def write_workbook(path, sheet_name="Sheet1", cells=None):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    for cell, value in (cells or {}).items():
        worksheet[cell] = value
    workbook.save(path)


@pytest.fixture()
def built_class():
    return build(
        Options(
            required_files=("submission.xlsx",),
            entries=("A2", "A5"),
            kwargs={"reference_file": "reference.xlsx"},
        )
    )


@pytest.fixture()
def built_instance(built_class):
    return built_class()


def test_build_class(built_class):
    assert issubclass(built_class, unittest.TestCase)


def test_build_class_name(built_class):
    assert built_class.__name__ == "TestDataSeriesMatchReference"


def test_instance_has_test_method(built_instance):
    assert hasattr(built_instance, "test_data_series_match_reference_0")


def test_doc_func_same_range_default(built_instance):
    assert (
        built_instance.test_data_series_match_reference_0.__doc__
        == "Check that the data series in range A2:A5 on sheet `<first worksheet>` exactly matches the reference values at the same cell locations."
    )


def test_doc_func_search_mode():
    built_class = build(
        Options(
            required_files=("submission.xlsx",),
            entries=("A2", "A5"),
            range_matches_reference=False,
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class()

    assert (
        built_instance.test_data_series_match_reference_0.__doc__
        == "Check that the data series in range A2:A5 on sheet `<first worksheet>` exactly matches somewhere in the submission workbook."
    )


def test_passing_same_range_case(fix_syspath):
    write_workbook(
        fix_syspath / "reference.xlsx",
        cells={"A2": 10, "A3": 20, "A4": 30, "A5": 40},
    )
    write_workbook(
        fix_syspath / "submission.xlsx",
        cells={"A2": 10, "A3": 20, "A4": 30, "A5": 40},
    )

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A2", "A5"),
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_data_series_match_reference_0")
    test_method = built_instance.test_data_series_match_reference_0
    test_method()

    assert test_method.__score__ == test_method.__weight__


def test_failing_same_range_case(fix_syspath):
    write_workbook(
        fix_syspath / "reference.xlsx",
        cells={"A2": 10, "A3": 20, "A4": 30, "A5": 40},
    )
    write_workbook(
        fix_syspath / "submission.xlsx",
        cells={"A2": 10, "A3": 20, "A4": 999, "A5": 40},
    )

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A2", "A5"),
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_data_series_match_reference_0")
    test_method = built_instance.test_data_series_match_reference_0

    with pytest.raises(AssertionError) as exc_info:
        test_method()

    assert "did not meet the expected" in str(exc_info.value)
    assert test_method.__score__ == 0


def test_passing_search_mode_relocated_case(fix_syspath):
    write_workbook(
        fix_syspath / "reference.xlsx",
        cells={"A2": 10, "A3": 20, "A4": 30, "A5": 40},
    )
    write_workbook(
        fix_syspath / "submission.xlsx",
        cells={"D2": 10, "D3": 20, "D4": 30, "D5": 40},
    )

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A2", "A5"),
            range_matches_reference=False,
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_data_series_match_reference_0")
    test_method = built_instance.test_data_series_match_reference_0
    test_method()

    assert test_method.__score__ == test_method.__weight__


def test_passing_search_mode_ratio_case(fix_syspath):
    write_workbook(
        fix_syspath / "reference.xlsx",
        cells={"A2": 10, "A3": 20, "A4": 30, "A5": 40},
    )
    write_workbook(
        fix_syspath / "submission.xlsx",
        cells={"D2": 10, "D3": 20, "D4": 999, "D5": 40},
    )

    built_class = build(
        Options(
            weight=1,
            ratio=0.75,
            required_files=("submission.xlsx",),
            entries=("A2", "A5"),
            range_matches_reference=False,
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_data_series_match_reference_0")
    test_method = built_instance.test_data_series_match_reference_0
    test_method()

    assert test_method.__score__ == test_method.__weight__


def test_failing_search_mode_exact_case(fix_syspath):
    write_workbook(
        fix_syspath / "reference.xlsx",
        cells={"A2": 10, "A3": 20, "A4": 30, "A5": 40},
    )
    write_workbook(
        fix_syspath / "submission.xlsx",
        cells={"D2": 10, "D3": 20, "D4": 999, "D5": 40},
    )

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A2", "A5"),
            range_matches_reference=False,
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_data_series_match_reference_0")
    test_method = built_instance.test_data_series_match_reference_0

    with pytest.raises(AssertionError) as exc_info:
        test_method()

    assert "No exact data series match" in str(exc_info.value)
    assert test_method.__score__ == 0


def test_first_worksheet_default_sheet_fallback(fix_syspath):
    write_workbook(
        fix_syspath / "reference.xlsx",
        sheet_name="Grades",
        cells={"A2": 10, "A3": 20, "A4": 30, "A5": 40},
    )
    write_workbook(
        fix_syspath / "submission.xlsx",
        sheet_name="Grades",
        cells={"A2": 10, "A3": 20, "A4": 30, "A5": 40},
    )

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A2", "A5"),
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_data_series_match_reference_0")
    test_method = built_instance.test_data_series_match_reference_0
    test_method()

    assert test_method.__score__ == test_method.__weight__


# ---------------------------------------------------------------------------
# `series_require_formulas` formula-detection gate: ordinary formula, array
# formula, plain value, and empty cell -- for BOTH the range_matches_reference
# branch (cell-by-cell) and the search-anywhere branch. Array-formula cells
# hold an openpyxl `ArrayFormula` object rather than a `str`, so a naive
# `isinstance(value, str)` check reports a legitimate array formula as "not
# a formula" -- this is the bug these tests pin down.
#
# Note: openpyxl-authored fixtures never carry a *cached* value for a
# formula cell (there's no calc engine involved), so a formula cell's
# `data_only` read is always None. The value-match `ratio` check and the
# formula-existence check are independent gates in the source, so the
# formula-passing tests below set `ratio=0.0` to isolate the formula gate
# from the (here, unrepresentable) value-match gate; the failing tests use
# values that already satisfy the default ratio so the formula gate is what
# trips the failure.
# ---------------------------------------------------------------------------


def test_require_formulas_same_range_ordinary_formula_passes(fix_syspath):
    write_workbook(fix_syspath / "reference.xlsx", cells={"A2": 10})
    write_workbook(fix_syspath / "submission.xlsx", cells={"A2": "=1+1"})

    built_class = build(
        Options(
            weight=1,
            ratio=0.0,
            required_files=("submission.xlsx",),
            entries=("A2", "A2"),
            series_require_formulas=True,
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_data_series_match_reference_0")
    test_method = built_instance.test_data_series_match_reference_0
    test_method()

    assert test_method.__score__ == test_method.__weight__


def test_require_formulas_same_range_array_formula_passes(fix_syspath):
    write_workbook(fix_syspath / "reference.xlsx", cells={"A2": 10})

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet["A2"] = ArrayFormula("A2", "=MODE.MULT(B1:B10)")
    workbook.save(fix_syspath / "submission.xlsx")

    built_class = build(
        Options(
            weight=1,
            ratio=0.0,
            required_files=("submission.xlsx",),
            entries=("A2", "A2"),
            series_require_formulas=True,
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_data_series_match_reference_0")
    test_method = built_instance.test_data_series_match_reference_0
    test_method()

    assert test_method.__score__ == test_method.__weight__


def test_require_formulas_same_range_plain_value_fails(fix_syspath):
    write_workbook(fix_syspath / "reference.xlsx", cells={"A2": 10})
    write_workbook(fix_syspath / "submission.xlsx", cells={"A2": 10})

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A2", "A2"),
            series_require_formulas=True,
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_data_series_match_reference_0")
    test_method = built_instance.test_data_series_match_reference_0

    with pytest.raises(AssertionError) as exc_info:
        test_method()

    assert "must contain a formula" in str(exc_info.value)
    assert test_method.__score__ == 0


def test_require_formulas_same_range_empty_cell_fails(fix_syspath):
    write_workbook(fix_syspath / "reference.xlsx", cells={})
    write_workbook(fix_syspath / "submission.xlsx", cells={})

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A2", "A2"),
            series_require_formulas=True,
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_data_series_match_reference_0")
    test_method = built_instance.test_data_series_match_reference_0

    with pytest.raises(AssertionError) as exc_info:
        test_method()

    assert "must contain a formula" in str(exc_info.value)
    assert test_method.__score__ == 0


def test_require_formulas_search_mode_ordinary_formula_passes(fix_syspath):
    write_workbook(fix_syspath / "reference.xlsx", cells={"A2": 10})
    write_workbook(fix_syspath / "submission.xlsx", cells={"A1": "=1+1"})

    built_class = build(
        Options(
            weight=1,
            ratio=0.0,
            required_files=("submission.xlsx",),
            entries=("A2", "A2"),
            range_matches_reference=False,
            series_require_formulas=True,
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_data_series_match_reference_0")
    test_method = built_instance.test_data_series_match_reference_0
    test_method()

    assert test_method.__score__ == test_method.__weight__


def test_require_formulas_search_mode_array_formula_passes(fix_syspath):
    write_workbook(fix_syspath / "reference.xlsx", cells={"A2": 10})

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet["A1"] = ArrayFormula("A1", "=MODE.MULT(B1:B10)")
    workbook.save(fix_syspath / "submission.xlsx")

    built_class = build(
        Options(
            weight=1,
            ratio=0.0,
            required_files=("submission.xlsx",),
            entries=("A2", "A2"),
            range_matches_reference=False,
            series_require_formulas=True,
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_data_series_match_reference_0")
    test_method = built_instance.test_data_series_match_reference_0
    test_method()

    assert test_method.__score__ == test_method.__weight__


def test_require_formulas_search_mode_plain_value_fails(fix_syspath):
    write_workbook(fix_syspath / "reference.xlsx", cells={"A2": 10})
    write_workbook(fix_syspath / "submission.xlsx", cells={"A1": 42})

    built_class = build(
        Options(
            weight=1,
            ratio=0.0,
            required_files=("submission.xlsx",),
            entries=("A2", "A2"),
            range_matches_reference=False,
            series_require_formulas=True,
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_data_series_match_reference_0")
    test_method = built_instance.test_data_series_match_reference_0

    with pytest.raises(AssertionError) as exc_info:
        test_method()

    assert "is not a formula cell" in str(exc_info.value)
    assert test_method.__score__ == 0


def test_require_formulas_search_mode_empty_cell_fails(fix_syspath):
    write_workbook(fix_syspath / "reference.xlsx", cells={"A2": 10})
    # A1 stays empty; B1 anchors sheet dimensions so the search space isn't
    # empty, while A1 (the first, and therefore best-scoring-tie-breaking,
    # length-1 candidate window) remains unfilled.
    write_workbook(fix_syspath / "submission.xlsx", cells={"B1": "unrelated"})

    built_class = build(
        Options(
            weight=1,
            ratio=0.0,
            required_files=("submission.xlsx",),
            entries=("A2", "A2"),
            range_matches_reference=False,
            series_require_formulas=True,
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_data_series_match_reference_0")
    test_method = built_instance.test_data_series_match_reference_0

    with pytest.raises(AssertionError) as exc_info:
        test_method()

    assert "is not a formula cell" in str(exc_info.value)
    assert test_method.__score__ == 0
