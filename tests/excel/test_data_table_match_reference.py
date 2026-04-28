import unittest

import pytest
from openpyxl import Workbook
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

from generic_grader.excel.data_table_match_reference import build
from generic_grader.utils.options import Options


def write_table_workbook(path, sheet_name="Sheet1", headers=None, data_rows=None, anchor="A1"):
    """Write a workbook with a table (header row + data rows) starting at *anchor*."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name

    col_letter, row_num = coordinate_from_string(anchor)
    start_row = row_num
    start_col = column_index_from_string(col_letter)

    headers = headers or []
    data_rows = data_rows or []

    for col_offset, header in enumerate(headers):
        worksheet.cell(start_row, start_col + col_offset, header)

    for row_offset, row in enumerate(data_rows, start=1):
        for col_offset, value in enumerate(row):
            worksheet.cell(start_row + row_offset, start_col + col_offset, value)

    workbook.save(path)


REF_HEADERS = ["Year", "Revenue", "Profit"]
REF_DATA = [[2020, 100, 20], [2021, 120, 25], [2022, 140, 30]]
END_CELL = "C4"  # A1:C4 => 1 header + 3 data rows


@pytest.fixture()
def built_class():
    return build(
        Options(
            required_files=("submission.xlsx",),
            entries=("A1", END_CELL),
            kwargs={"reference_file": "reference.xlsx"},
        )
    )


@pytest.fixture()
def built_instance(built_class):
    return built_class()


# ------------------------------------------------------------------
# Class / instance shape
# ------------------------------------------------------------------


def test_build_class(built_class):
    assert issubclass(built_class, unittest.TestCase)


def test_build_class_name(built_class):
    assert built_class.__name__ == "TestDataTableMatchReference"


def test_instance_has_test_method(built_instance):
    assert hasattr(built_instance, "test_data_table_match_reference_0")


# ------------------------------------------------------------------
# Doc strings
# ------------------------------------------------------------------


def test_doc_func_range_mode(built_instance):
    assert built_instance.test_data_table_match_reference_0.__doc__ == (
        f"Check that the data table in range A1:{END_CELL}"
        " on sheet `<first worksheet>` matches at the same location as the reference table,"
        " with columns matched by header name."
    )


def test_doc_func_search_mode():
    built_class = build(
        Options(
            required_files=("submission.xlsx",),
            entries=("A1", END_CELL),
            range_matches_reference=False,
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class()
    assert built_instance.test_data_table_match_reference_0.__doc__ == (
        f"Check that the data table in range A1:{END_CELL}"
        " on sheet `<first worksheet>` matches anywhere in the submission matching the reference table,"
        " with columns matched by header name."
    )


# ------------------------------------------------------------------
# Passing cases
# ------------------------------------------------------------------


def test_passing_exact_same_location(fix_syspath):
    """Same headers, same location, same data — should pass."""
    write_table_workbook(fix_syspath / "reference.xlsx", headers=REF_HEADERS, data_rows=REF_DATA)
    write_table_workbook(fix_syspath / "submission.xlsx", headers=REF_HEADERS, data_rows=REF_DATA)

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A1", END_CELL),
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_data_table_match_reference_0")
    test_method = built_instance.test_data_table_match_reference_0
    test_method()
    assert test_method.__score__ == test_method.__weight__


def test_passing_fuzzy_header_case_difference(fix_syspath):
    """Headers differ only in case — should pass with default fuzzy matching."""
    sub_headers = ["year", "revenue", "profit"]
    write_table_workbook(fix_syspath / "reference.xlsx", headers=REF_HEADERS, data_rows=REF_DATA)
    write_table_workbook(fix_syspath / "submission.xlsx", headers=sub_headers, data_rows=REF_DATA)

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A1", END_CELL),
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_data_table_match_reference_0")
    test_method = built_instance.test_data_table_match_reference_0
    test_method()
    assert test_method.__score__ == test_method.__weight__


def test_passing_reordered_columns(fix_syspath):
    """Columns in different order — should pass since column order is flexible by default."""
    sub_headers = ["Profit", "Year", "Revenue"]
    # Rearrange data to match reordered headers
    sub_data = [[row[2], row[0], row[1]] for row in REF_DATA]

    write_table_workbook(fix_syspath / "reference.xlsx", headers=REF_HEADERS, data_rows=REF_DATA)
    write_table_workbook(fix_syspath / "submission.xlsx", headers=sub_headers, data_rows=sub_data)

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A1", END_CELL),
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_data_table_match_reference_0")
    test_method = built_instance.test_data_table_match_reference_0
    test_method()
    assert test_method.__score__ == test_method.__weight__


def test_passing_relocated_table_search_mode(fix_syspath):
    """Table is at a different location in the submission — search mode finds it."""
    write_table_workbook(fix_syspath / "reference.xlsx", headers=REF_HEADERS, data_rows=REF_DATA)
    write_table_workbook(
        fix_syspath / "submission.xlsx",
        headers=REF_HEADERS,
        data_rows=REF_DATA,
        anchor="D5",
    )

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A1", END_CELL),
            range_matches_reference=False,
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_data_table_match_reference_0")
    test_method = built_instance.test_data_table_match_reference_0
    test_method()
    assert test_method.__score__ == test_method.__weight__


def test_passing_fuzzy_relocated_search_mode(fix_syspath):
    """Fuzzy headers at a different location — search mode + fuzzy headers both apply."""
    sub_headers = ["year", "revenue", "profit"]
    write_table_workbook(fix_syspath / "reference.xlsx", headers=REF_HEADERS, data_rows=REF_DATA)
    write_table_workbook(
        fix_syspath / "submission.xlsx",
        headers=sub_headers,
        data_rows=REF_DATA,
        anchor="B3",
    )

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A1", END_CELL),
            range_matches_reference=False,
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_data_table_match_reference_0")
    test_method = built_instance.test_data_table_match_reference_0
    test_method()
    assert test_method.__score__ == test_method.__weight__


# ------------------------------------------------------------------
# Failing cases
# ------------------------------------------------------------------


def test_failing_wrong_data_values(fix_syspath):
    """Data values don't match — should fail with a data mismatch hint."""
    bad_data = [[2020, 100, 999], [2021, 120, 25], [2022, 140, 30]]
    write_table_workbook(fix_syspath / "reference.xlsx", headers=REF_HEADERS, data_rows=REF_DATA)
    write_table_workbook(fix_syspath / "submission.xlsx", headers=REF_HEADERS, data_rows=bad_data)

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A1", END_CELL),
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_data_table_match_reference_0")
    test_method = built_instance.test_data_table_match_reference_0
    with pytest.raises(AssertionError) as exc_info:
        test_method()
    assert "Data mismatch" in str(exc_info.value)
    assert test_method.__score__ == 0


def test_failing_unrecognized_headers(fix_syspath):
    """Headers are completely different — should fail header matching."""
    bad_headers = ["Alpha", "Beta", "Gamma"]
    write_table_workbook(fix_syspath / "reference.xlsx", headers=REF_HEADERS, data_rows=REF_DATA)
    write_table_workbook(
        fix_syspath / "submission.xlsx", headers=bad_headers, data_rows=REF_DATA
    )

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A1", END_CELL),
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_data_table_match_reference_0")
    test_method = built_instance.test_data_table_match_reference_0
    with pytest.raises(AssertionError) as exc_info:
        test_method()
    assert "Could not match" in str(exc_info.value)
    assert test_method.__score__ == 0


def test_failing_table_not_found_search_mode(fix_syspath):
    """Table not found anywhere in search mode — should fail with clear hint."""
    write_table_workbook(fix_syspath / "reference.xlsx", headers=REF_HEADERS, data_rows=REF_DATA)
    write_table_workbook(
        fix_syspath / "submission.xlsx",
        headers=["Alpha", "Beta", "Gamma"],
        data_rows=REF_DATA,
    )

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A1", END_CELL),
            range_matches_reference=False,
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_data_table_match_reference_0")
    test_method = built_instance.test_data_table_match_reference_0
    with pytest.raises(AssertionError) as exc_info:
        test_method()
    assert "Could not find a table" in str(exc_info.value)
    assert test_method.__score__ == 0


def test_failing_strict_headers_rejects_fuzzy(fix_syspath):
    """strict_headers=True should reject case-differing headers."""
    sub_headers = ["year", "revenue", "profit"]
    write_table_workbook(fix_syspath / "reference.xlsx", headers=REF_HEADERS, data_rows=REF_DATA)
    write_table_workbook(fix_syspath / "submission.xlsx", headers=sub_headers, data_rows=REF_DATA)

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A1", END_CELL),
            strict_headers=True,
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_data_table_match_reference_0")
    test_method = built_instance.test_data_table_match_reference_0
    with pytest.raises(AssertionError) as exc_info:
        test_method()
    assert "Could not match" in str(exc_info.value)
    assert test_method.__score__ == 0


def test_failing_strict_column_order_rejects_reordered(fix_syspath):
    """strict_column_order=True should reject columns in a different order."""
    sub_headers = ["Profit", "Year", "Revenue"]
    sub_data = [[row[2], row[0], row[1]] for row in REF_DATA]

    write_table_workbook(fix_syspath / "reference.xlsx", headers=REF_HEADERS, data_rows=REF_DATA)
    write_table_workbook(fix_syspath / "submission.xlsx", headers=sub_headers, data_rows=sub_data)

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A1", END_CELL),
            strict_column_order=True,
            kwargs={"reference_file": "reference.xlsx"},
        )
    )
    built_instance = built_class(methodName="test_data_table_match_reference_0")
    test_method = built_instance.test_data_table_match_reference_0
    with pytest.raises(AssertionError) as exc_info:
        test_method()
    assert "Could not match" in str(exc_info.value)
    assert test_method.__score__ == 0
