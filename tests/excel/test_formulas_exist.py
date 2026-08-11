import unittest

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.formula import ArrayFormula

from generic_grader.excel.formulas_exist import build
from generic_grader.utils.options import Options


def write_workbook(path, sheet_name="Sheet1", cells=None):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    for cell, value in (cells or {}).items():
        worksheet[cell] = value
    workbook.save(path)


@pytest.fixture()
def built_class():
    return build(
        Options(
            required_files=("submission.xlsx",),
            entries=("A1", "B1"),
        )
    )


@pytest.fixture()
def built_instance(built_class):
    return built_class()


def test_build_class(built_class):
    assert issubclass(built_class, unittest.TestCase)


def test_build_class_name(built_class):
    assert built_class.__name__ == "TestFormulasExist"


def test_instance_has_test_method(built_instance):
    assert hasattr(built_instance, "test_formulas_exist_0")


def test_doc_func_same_range_default(built_instance):
    assert (
        built_instance.test_formulas_exist_0.__doc__
        == "Check that cells in range A1:B1 on sheet `<first worksheet>` use formulas."
    )


def test_doc_func_search_mode():
    built_class = build(
        Options(
            required_files=("submission.xlsx",),
            entries=("A1", "B1"),
            range_matches_reference=False,
        )
    )
    built_instance = built_class()

    assert (
        built_instance.test_formulas_exist_0.__doc__
        == "Check that cells in range A1:B1 on sheet `<first worksheet>` have a same-sized formula-only region somewhere in the submission workbook."
    )


def test_passing_same_range_case(fix_syspath):
    write_workbook(
        fix_syspath / "submission.xlsx",
        cells={"A1": "=1+1", "B1": "=2+2"},
    )

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A1", "B1"),
        )
    )
    built_instance = built_class(methodName="test_formulas_exist_0")
    test_method = built_instance.test_formulas_exist_0
    test_method()

    assert test_method.__score__ == test_method.__weight__


def test_failing_same_range_case(fix_syspath):
    write_workbook(
        fix_syspath / "submission.xlsx",
        cells={"A1": "=1+1", "B1": 4},
    )

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A1", "B1"),
        )
    )
    built_instance = built_class(methodName="test_formulas_exist_0")
    test_method = built_instance.test_formulas_exist_0

    with pytest.raises(AssertionError) as exc_info:
        test_method()

    assert "Cell B1" in str(exc_info.value)
    assert test_method.__score__ == 0


def test_passing_search_whole_sheet_case(fix_syspath):
    write_workbook(
        fix_syspath / "submission.xlsx",
        cells={"D4": "=1+1", "E4": "=2+2"},
    )

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A1", "B1"),
            range_matches_reference=False,
        )
    )
    built_instance = built_class(methodName="test_formulas_exist_0")
    test_method = built_instance.test_formulas_exist_0
    test_method()

    assert test_method.__score__ == test_method.__weight__


def test_failing_search_whole_sheet_case(fix_syspath):
    write_workbook(
        fix_syspath / "submission.xlsx",
        cells={"D4": "=1+1", "E4": 2},
    )

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A1", "B1"),
            range_matches_reference=False,
        )
    )
    built_instance = built_class(methodName="test_formulas_exist_0")
    test_method = built_instance.test_formulas_exist_0

    with pytest.raises(AssertionError) as exc_info:
        test_method()

    assert "formula-only region" in str(exc_info.value)
    assert test_method.__score__ == 0


def test_first_worksheet_default_sheet_fallback(fix_syspath):
    write_workbook(
        fix_syspath / "submission.xlsx",
        sheet_name="Grades",
        cells={"A1": "=1+1", "B1": "=2+2"},
    )

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A1", "B1"),
        )
    )
    built_instance = built_class(methodName="test_formulas_exist_0")
    test_method = built_instance.test_formulas_exist_0
    test_method()

    assert test_method.__score__ == test_method.__weight__


# ---------------------------------------------------------------------------
# Per-cell formula-detection gate: ordinary formula, array formula, plain
# value, and empty cell. Array-formula cells hold an openpyxl `ArrayFormula`
# object rather than a `str`, so a naive `isinstance(value, str)` check
# reports a legitimate array formula as "not a formula" -- this is the bug
# these tests pin down.
# ---------------------------------------------------------------------------


def test_ordinary_string_formula_is_detected(fix_syspath):
    write_workbook(
        fix_syspath / "submission.xlsx",
        cells={"A1": "=1+1"},
    )

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A1", "A1"),
        )
    )
    built_instance = built_class(methodName="test_formulas_exist_0")
    test_method = built_instance.test_formulas_exist_0
    test_method()

    assert test_method.__score__ == test_method.__weight__


def test_array_formula_is_detected(fix_syspath):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet["A1"] = ArrayFormula("A1", "=MODE.MULT(B1:B10)")
    workbook.save(fix_syspath / "submission.xlsx")

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A1", "A1"),
        )
    )
    built_instance = built_class(methodName="test_formulas_exist_0")
    test_method = built_instance.test_formulas_exist_0
    test_method()

    assert test_method.__score__ == test_method.__weight__


def test_plain_value_is_not_a_formula(fix_syspath):
    write_workbook(
        fix_syspath / "submission.xlsx",
        cells={"A1": 42},
    )

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A1", "A1"),
        )
    )
    built_instance = built_class(methodName="test_formulas_exist_0")
    test_method = built_instance.test_formulas_exist_0

    with pytest.raises(AssertionError) as exc_info:
        test_method()

    assert "Cell A1" in str(exc_info.value)
    assert test_method.__score__ == 0


def test_empty_cell_is_not_a_formula(fix_syspath):
    write_workbook(
        fix_syspath / "submission.xlsx",
        cells={},
    )

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A1", "A1"),
        )
    )
    built_instance = built_class(methodName="test_formulas_exist_0")
    test_method = built_instance.test_formulas_exist_0

    with pytest.raises(AssertionError) as exc_info:
        test_method()

    assert "Cell A1" in str(exc_info.value)
    assert test_method.__score__ == 0


def test_array_formula_with_no_text_is_still_detected(fix_syspath):
    """An ArrayFormula with text=None is still a formula, not a plain value.

    openpyxl's ArrayFormula.__init__ defaults text=None (e.g. for non-anchor
    cells of a multi-cell array range); is_formula_value must not treat that
    as "not a formula".
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet["A1"] = ArrayFormula("A1")
    workbook.save(fix_syspath / "submission.xlsx")

    built_class = build(
        Options(
            weight=1,
            required_files=("submission.xlsx",),
            entries=("A1", "A1"),
        )
    )
    built_instance = built_class(methodName="test_formulas_exist_0")
    test_method = built_instance.test_formulas_exist_0
    test_method()

    assert test_method.__score__ == test_method.__weight__
