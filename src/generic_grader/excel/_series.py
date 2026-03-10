"""Shared helpers for Excel data series checks."""

from math import isclose

from openpyxl.utils.cell import get_column_letter, range_boundaries


def _normalize_value(value):
    if isinstance(value, str):
        return value.strip().casefold()
    return value


def _values_match(a, b, relative_tolerance: float, absolute_tolerance: float) -> bool:
    a_value = _normalize_value(a)
    b_value = _normalize_value(b)

    if isinstance(a_value, bool) or isinstance(b_value, bool):
        return a_value == b_value

    if isinstance(a_value, (int, float)) and isinstance(b_value, (int, float)):
        return isclose(
            float(a_value),
            float(b_value),
            rel_tol=relative_tolerance,
            abs_tol=absolute_tolerance,
        )

    return a_value == b_value


def _series_ratio(
    candidate: list,
    target: list,
    relative_tolerance: float,
    absolute_tolerance: float,
) -> float:
    if len(candidate) != len(target):
        return 0.0

    if not target:
        return 1.0

    n_matches = sum(
        _values_match(c, t, relative_tolerance, absolute_tolerance)
        for c, t in zip(candidate, target, strict=True)
    )
    return n_matches / len(target)


def _coord(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"


def normalize_search_orientation(search_orientation: str) -> tuple[str, ...]:
    orientation = search_orientation.strip().lower()
    if orientation in {"either", "both"}:
        return ("row", "column")
    if orientation in {"row", "rows", "horizontal"}:
        return ("row",)
    if orientation in {"column", "columns", "col", "vertical"}:
        return ("column",)
    raise ValueError(
        "`kwargs['search_orientation']` must be one of"
        " 'either', 'row', or 'column'."
    )


def extract_series_from_range(sheet, start_cell: str, end_cell: str) -> dict:
    """Extract a single row/column series from a sheet range."""

    min_col, min_row, max_col, max_row = range_boundaries(f"{start_cell}:{end_cell}")

    if min_row != max_row and min_col != max_col:
        raise ValueError(
            "Data series checks require a single row or single column range"
            " for `entries`, e.g. ('A2', 'A10') or ('B4', 'H4')."
        )

    if min_row == max_row:
        orientation = "row"
        values = [sheet.cell(min_row, col).value for col in range(min_col, max_col + 1)]
    else:
        orientation = "column"
        values = [sheet.cell(row, min_col).value for row in range(min_row, max_row + 1)]

    return {
        "orientation": orientation,
        "values": values,
        "start_cell": _coord(min_row, min_col),
        "end_cell": _coord(max_row, max_col),
    }


def find_best_series_window(
    sheet,
    target: list,
    search_orientation: str,
    relative_tolerance: float,
    absolute_tolerance: float,
):
    """Return the best matching row/column window for target values."""

    if not target:
        return None

    max_row = sheet.max_row
    max_col = sheet.max_column
    if max_row == 0 or max_col == 0:
        return None

    length = len(target)
    orientations = normalize_search_orientation(search_orientation)

    best_match = None
    best_ratio = -1.0

    if "row" in orientations and length <= max_col:
        for row in range(1, max_row + 1):
            for col in range(1, max_col - length + 2):
                candidate = [sheet.cell(row, col + i).value for i in range(length)]
                ratio = _series_ratio(
                    candidate,
                    target,
                    relative_tolerance,
                    absolute_tolerance,
                )
                if ratio > best_ratio:
                    best_ratio = ratio
                    best_match = {
                        "orientation": "row",
                        "start_cell": _coord(row, col),
                        "end_cell": _coord(row, col + length - 1),
                        "ratio": ratio,
                    }

    if "column" in orientations and length <= max_row:
        for col in range(1, max_col + 1):
            for row in range(1, max_row - length + 2):
                candidate = [sheet.cell(row + i, col).value for i in range(length)]
                ratio = _series_ratio(
                    candidate,
                    target,
                    relative_tolerance,
                    absolute_tolerance,
                )
                if ratio > best_ratio:
                    best_ratio = ratio
                    best_match = {
                        "orientation": "column",
                        "start_cell": _coord(row, col),
                        "end_cell": _coord(row + length - 1, col),
                        "ratio": ratio,
                    }

    return best_match


def find_exact_series_window(
    sheet,
    target: list,
    search_orientation: str,
    relative_tolerance: float,
    absolute_tolerance: float,
):
    """Return a matching row/column window only when all values match."""

    best = find_best_series_window(
        sheet,
        target,
        search_orientation,
        relative_tolerance,
        absolute_tolerance,
    )
    if best is None:
        return None
    if best["ratio"] >= 1.0:
        return best
    return None
